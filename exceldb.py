from openpyxl import Workbook, load_workbook
import os



def indexvalue(ws):
	count = 0
	for ID in ws['B']:
		if ID.value != None:
			count += 1

	index = count + 1
	return (index)

def savedoc(wb):
	wb.save('/home/HangTuah/mysite/exceldb.xlsx')

def updatecell(ws , number , index):
	cell = 'B' + str(index)
	ws[cell] = number

def retrieveID(ws, index):
	cell = 'A' + str(index)
	return ws[cell].value

def checknumber(ws, number):
	for num in ws['B']:
		if num.value != None:
			if num.value == number:
				return 0
	return 100

