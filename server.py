#flask imports
from flask import Flask, render_template, request, redirect, url_for, session
#gspread imports
from google.oauth2 import service_account
from googleapiclient.discovery import build
#Flaskform imports
from flask_wtf import FlaskForm
#wtforms imports
from wtforms import IntegerField
from wtforms.validators import InputRequired, ValidationError
#custom files
from phoneverification import OTPVerification
#sql
from flask_sqlalchemy import SQLAlchemy
#jinja protection
from jinja2 import Environment, select_autoescape
from exceldb import *
from openpyxl import Workbook, load_workbook
import os

env = Environment(autoescape=select_autoescape(
    enabled_extensions=('html', 'xml'),
    default_for_string=True,
))

#instantiating OTP
otp = OTPVerification()

#checks if number has 6 in front of it
#If not adds 6 to the number
def sixcheck(form, field):
    num = str(field.data)
    if num[0] != "6":
        num = '60' + str(field.data)
        field.data = int(num)

#checks length of phone number
def phonelengthcheck(form,field):
    num = str(field.data)
    if len(num) < 10 or len(num) > 13:
        raise ValidationError('Invalid phone number')

#checks whether the length of the OTP entered is 6
def OTPlength(form,field):
    otp = str(field.data)
    if len(otp) != 6:
        raise ValidationError('Invalid OTP')

#Checks whether OTP is equal to OTP in phoneverification.py
def OTP_validation(form,field):
    result = otp.check_otp(field.data)
    if result != 'success':
        raise ValidationError('Wrong OTP')
    if result == 'timeout':
        raise ValidationError('OTP Timed Out, Please Resend the OTP')
    if result == 'already':
        raise ValidationError('OTP already claimed, no longer valid.')

#custom Functions
#Set Session Variables
#if 1 yes if 0 no
def setvar():
#phone number used in this session
    session['number'] = 0
#Whether game IDs must be displayed
    session['displayid'] = 0
#if angpao must be displayed
    session['displayclaim'] = 0
#if user requested otp
    session['OTPasked'] = 0
#should user be allowed to load the previous page
    session['noreturn'] = 0
    session['pwidlaunch'] = 0


#Clear all session variables except noreturn which is cleared on landing page
def clearvar():
    session.pop('number', None)
    session.pop('OTPasked', None)
    session.pop('displayid', None)
    session.pop('displayclaim', None)

#checks whether 3 minutes have passed before allowing to reload numverif
def check_timer():
    session['OTPasked'] = otp.checktime()


#FORM CLASSES
#PhoneInput
class PhoneNumber(FlaskForm):
#Phone Number Field With validators
    phone = IntegerField('Enter Phone Number:', validators=[
        InputRequired(), sixcheck, phonelengthcheck])

class OTP_form(FlaskForm):
#OTPInput with validators
    OTP = IntegerField('Enter OTP:', validators=[
        InputRequired(), OTPlength, OTP_validation])

###################################################################################
#App definition
#Secret Key Init

app = Flask(__name__)
app.config['SECRET_KEY'] = 'HangTuah2021LeblakcRenditionStructure'

###################################################################################
#Views and URLs for the website
@app.route('/', methods=['GET','POST'])
def numverif():
#100 is returned if clock uninitialised meaning the session has just started/re-started
#sets session variables
    if otp.checktime() == 100:
            setvar()
#Instantiate form
    form = PhoneNumber()
#security verification of noreturn variable
    session['noreturn'] = 0
#if request is post & form validated
# assign session number variable
# Verify if number has already claimed
    if request.method == 'POST' and form.validate_on_submit():
        session['number'] = form.phone.data
        ver = vernum()
#If number unclaimed generate OTP
#set OTPasked to 1: Meaning OTP timer has started
#direct to OTP page
        if ver != 0:
            otp.generate_otp(session['number'])
            session['OTPasked'] = 1
            return redirect(url_for('otp_input'))
#if attempting to reload this page, check timer first
    if session['OTPasked'] == 1:
        check_timer()
#when 3 minutes has passed, OTPasked will reset to 0 and allow page to load
    if session['OTPasked'] == 0:
        print("Numverif")
        return render_template('numverif.html', form=form)
    else:
        return redirect(url_for('otp_input'))

@app.route('/OTP',methods=['GET','POST'])
def otp_input():
    if session['noreturn'] == 1:
        return redirect(url_for('numverif'))
#if timer is on display OTP form
#else redirect to numverif
    if session['OTPasked'] == 1:
        otpform = OTP_form()
#if post + validated, allow angpao to be displayed by setting variable
#redirect to claim
        if request.method == 'POST' and otpform.validate_on_submit():
            session['displayclaim'] = 1
            return redirect(url_for('claim'))
        return render_template('otpinput.html', form=otpform)
    else:
        return redirect(url_for('numverif'))

@app.route('/claim', methods=['GET','POST'])
def claim():
    if session['noreturn'] == 1:
        return redirect(url_for('numverif'))
#if allowed, display angpaoclaim
    if session['displayclaim'] == 1:
#if post disable OTP asked
#Allow Display of ID/Final page
#redirect to final page
        if request.method=='POST':
            session['OTPasked'] = 0
            session['displayid'] = 1
            session['pwidlaunch'] = 1
            return redirect(url_for('final'))

        return render_template('angpaoclaim.html')
    else:
        return redirect(url_for('numverif'))

@app.route('/final', methods=['GET'])
def final():
    if session['noreturn'] == 1:
        return redirect(url_for('numverif'))
#if allowed, display ID
#clear variables and set no return to 1 to prevent backpedalling and re-claiming
    if session['displayid'] == 1:
        if session['pwidlaunch'] == 1:
            ID = pwid()
            ID = "0" + str(ID)
        clearvar()
        session['noreturn'] = 1
        return render_template('final.html',id = ID, pw="Aaa888")
    else:
        return redirect(url_for('numverif'))

#####################################################################################
if __name__ == '__main__':
    app.run(debug=True)
    app.config.update(
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE='Lax',
    )

#####################################################################################
def vernum():
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    my_file = os.path.join(THIS_FOLDER, 'exceldb.xlsx')
    wb = load_workbook(my_file)
    ws = wb.active

    num = session['number']
    check = checknumber(ws, num)
    wb.close()
    if check == 0:
        return 0
    else:
        return 100


#######################################################################

def pwid():
    if session['number']!=0:
        if session['pwidlaunch'] == 1:
            THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
            my_file = os.path.join(THIS_FOLDER, 'exceldb.xlsx')
            wb = load_workbook(my_file)
            ws = wb.active
            num = session['number']
            check = checknumber(ws, num)
            if check != 0:
                number = session['number']
                index = indexvalue(ws)
                updatecell(ws, number, index)
                ID = retrieveID(ws, index)
                savedoc(wb)
            session['pwidlaunch'] = 0
            wb.close()
            return ID
        return 0
    return 0